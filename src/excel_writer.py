"""
excel_writer.py — Creates / updates the bookshelf Excel catalog.

Columns: real_name | status | name | author | description | file_type | isbn_10 | isbn_13 | reason_for_failure
"""

from pathlib import Path
from typing import List, Dict, Set
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = [
    "real_name",
    "status",
    "name",
    "author",
    "description",
    "file_type",
    "isbn_10",
    "isbn_13",
    "reason_for_failure",
]
COL_WIDTHS = [40, 10, 40, 30, 60, 12, 18, 18, 50]

_HEADER_FILL = PatternFill("solid", start_color="1F3864")
_FAIL_FILL   = PatternFill("solid", start_color="FFE0E0")
_ALT_FILL    = PatternFill("solid", start_color="EBF0FA")
_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def load_processed_files(output_path: str) -> Set[str]:
    """
    Load the set of files already processed from an existing Excel file.
    Returns a set of 'real_name' values that have been processed.
    If file doesn't exist, returns empty set.
    """
    if not Path(output_path).exists():
        return set()
    
    try:
        wb = load_workbook(output_path)
        ws = wb.active
        processed = set()
        
        # Skip header row, read all real_name values (column 1)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            cell_value = row[0].value
            if cell_value and isinstance(cell_value, str) and not cell_value.startswith("Total:"):
                processed.add(cell_value)
        
        return processed
    except Exception as e:
        print(f"⚠️  Could not read existing Excel file: {e}")
        return set()


def _style_header(ws):
    for col_idx, (col_name, width) in enumerate(zip(COLUMNS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20


def _style_row(ws, row_idx: int, has_failure: bool):
    fill = _FAIL_FILL if has_failure else (_ALT_FILL if row_idx % 2 == 0 else None)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = _BORDER
        if fill:
            cell.fill = fill


def append_record_to_excel(record: Dict, output_path: str):
    """
    Append a single record to an existing Excel file.
    If file doesn't exist, create it with headers first.
    """
    output = Path(output_path)
    
    # Create or load workbook
    if output.exists():
        wb = load_workbook(output_path)
        ws = wb.active
        next_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Book Catalog"
        ws.freeze_panes = "A2"
        _style_header(ws)
        next_row = 2
    
    # Add the new record
    has_failure = bool(record.get("reason_for_failure", "").strip())
    ws.cell(row=next_row, column=1, value=record.get("real_name", ""))
    ws.cell(row=next_row, column=2, value=record.get("status", ""))
    ws.cell(row=next_row, column=3, value=record.get("name", ""))
    ws.cell(row=next_row, column=4, value=record.get("author", ""))
    ws.cell(row=next_row, column=5, value=record.get("description", ""))
    ws.cell(row=next_row, column=6, value=record.get("file_type", ""))
    ws.cell(row=next_row, column=7, value=record.get("isbn_10", ""))
    ws.cell(row=next_row, column=8, value=record.get("isbn_13", ""))
    ws.cell(row=next_row, column=9, value=record.get("reason_for_failure", ""))
    _style_row(ws, next_row, has_failure)
    
    # Update summary row
    total = next_row - 1  # Exclude header
    ok = sum(1 for row in ws.iter_rows(min_row=2, max_row=next_row-1, min_col=9, max_col=9)
             if not (row[0].value and str(row[0].value).strip()))
    
    summary_row = next_row + 1
    ws.cell(row=summary_row, column=1, value=f"Total: {total}  |  Resolved: {ok}  |  Failed: {total - ok}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True, italic=True, name="Arial", size=10)
    
    wb.save(output_path)


def write_excel(records: List[Dict], output_path: str):
    """
    Write `records` to `output_path`.
    Each record should contain: {real_name, status, name, author, description, file_type, isbn_10, isbn_13, reason_for_failure}
    Overwrites if file exists.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Book Catalog"
    ws.freeze_panes = "A2"

    _style_header(ws)

    for i, rec in enumerate(records, start=2):
        has_failure = bool(rec.get("reason_for_failure", "").strip())
        ws.cell(row=i, column=1, value=rec.get("real_name", ""))
        ws.cell(row=i, column=2, value=rec.get("status", ""))
        ws.cell(row=i, column=3, value=rec.get("name", ""))
        ws.cell(row=i, column=4, value=rec.get("author", ""))
        ws.cell(row=i, column=5, value=rec.get("description", ""))
        ws.cell(row=i, column=6, value=rec.get("file_type", ""))
        ws.cell(row=i, column=7, value=rec.get("isbn_10", ""))
        ws.cell(row=i, column=8, value=rec.get("isbn_13", ""))
        ws.cell(row=i, column=9, value=rec.get("reason_for_failure", ""))
        _style_row(ws, i, has_failure)

    # Summary row
    total = len(records)
    ok = sum(1 for r in records if not r.get("reason_for_failure", "").strip())
    ws.cell(row=total + 3, column=1, value=f"Total: {total}  |  Resolved: {ok}  |  Failed: {total - ok}")
    ws.cell(row=total + 3, column=1).font = Font(bold=True, italic=True, name="Arial", size=10)

    wb.save(output_path)


def append_record_to_excel(record: Dict, output_path: str):
    """
    Append a single record to an existing Excel file.
    If file doesn't exist, create it with headers first.
    """
    output = Path(output_path)
    
    # Create or load workbook
    if output.exists():
        wb = load_workbook(output_path)
        ws = wb.active
        next_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Book Catalog"
        ws.freeze_panes = "A2"
        _style_header(ws)
        next_row = 2
    
    # Add the new record
    has_failure = bool(record.get("reason_for_failure", "").strip())
    ws.cell(row=next_row, column=1, value=record.get("real_name", ""))
    ws.cell(row=next_row, column=2, value=record.get("status", ""))
    ws.cell(row=next_row, column=3, value=record.get("name", ""))
    ws.cell(row=next_row, column=4, value=record.get("author", ""))
    ws.cell(row=next_row, column=5, value=record.get("description", ""))
    ws.cell(row=next_row, column=6, value=record.get("file_type", ""))
    ws.cell(row=next_row, column=7, value=record.get("isbn_10", ""))
    ws.cell(row=next_row, column=8, value=record.get("isbn_13", ""))
    ws.cell(row=next_row, column=9, value=record.get("reason_for_failure", ""))
    _style_row(ws, next_row, has_failure)
    
    # Update summary row
    total = next_row - 1  # Exclude header
    ok = sum(1 for row in ws.iter_rows(min_row=2, max_row=next_row-1, min_col=9, max_col=9)
             if not (row[0].value and str(row[0].value).strip()))
    
    summary_row = next_row + 1
    ws.cell(row=summary_row, column=1, value=f"Total: {total}  |  Resolved: {ok}  |  Failed: {total - ok}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True, italic=True, name="Arial", size=10)
    
    wb.save(output_path)


def write_excel(records: List[Dict], output_path: str):
    """
    Write `records` to `output_path`.
    Each record should contain: {real_name, status, name, author, description, file_type, isbn_10, isbn_13, reason_for_failure}
    Overwrites if file exists.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Book Catalog"
    ws.freeze_panes = "A2"

    _style_header(ws)

    for i, rec in enumerate(records, start=2):
        has_failure = bool(rec.get("reason_for_failure", "").strip())
        ws.cell(row=i, column=1, value=rec.get("real_name", ""))
        ws.cell(row=i, column=2, value=rec.get("status", ""))
        ws.cell(row=i, column=3, value=rec.get("name", ""))
        ws.cell(row=i, column=4, value=rec.get("author", ""))
        ws.cell(row=i, column=5, value=rec.get("description", ""))
        ws.cell(row=i, column=6, value=rec.get("file_type", ""))
        ws.cell(row=i, column=7, value=rec.get("isbn_10", ""))
        ws.cell(row=i, column=8, value=rec.get("isbn_13", ""))
        ws.cell(row=i, column=9, value=rec.get("reason_for_failure", ""))
        _style_row(ws, i, has_failure)

    # Summary row
    total = len(records)
    ok = sum(1 for r in records if not r.get("reason_for_failure", "").strip())
    ws.cell(row=total + 3, column=1, value=f"Total: {total}  |  Resolved: {ok}  |  Failed: {total - ok}")
    ws.cell(row=total + 3, column=1).font = Font(bold=True, italic=True, name="Arial", size=10)

    wb.save(output_path)
    print(f"[OK] Excel saved → {output_path}")
